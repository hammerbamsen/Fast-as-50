"""Excel-master for Fast as Fifty — erstatter Word-snapshottet for Kennet.

REGEL: Denne fil GENERERES fra data/plan.json + data.json. Ret den ALDRIG i hånden.
Master_Plan.xlsx blev afskaffet 16/7-2026 præcis fordi den blev vedligeholdt manuelt
og drev fra plan.json. Excel er backup og tracking — plan.json er master.

Faneblade:
  Oversigt    nu vs. mål (vægt, fedt, FTP, W/kg, CTL) + racedatoer
  Plan 2026   14-ugers Médoc-plan
  Plan 2027   51-ugers TdS-plan med FTP- og W/kg-mål pr. uge
  Tracking    daglig historik: vægt, fedt, CTL, TSB, HRV, søvn
  Ugesummer   planlagt vs. faktisk TSS pr. uge
  Zoner       aktuelle løbe- og cykelzoner
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Fast as Fifty-paletten (samme bordeaux/guld som dashboardet)
BORDEAUX = "4A1520"
GOLD = "C89B3C"
CREAM = "F2EDE4"
MUTED = "7A6A58"

H_FILL = PatternFill("solid", fgColor=BORDEAUX)
H_FONT = Font(color="FFFFFF", bold=True, size=10)
SUB_FILL = PatternFill("solid", fgColor=CREAM)
TITLE_FONT = Font(color=BORDEAUX, bold=True, size=14)
GOLD_FONT = Font(color=GOLD, bold=True, size=11)
THIN = Side(style="thin", color="DDD5C8")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _headers(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill, c.font, c.border = H_FILL, H_FONT, BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _row(ws, row, values, number_formats=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BOX
        c.alignment = Alignment(vertical="top", wrap_text=(i == len(values)))
        if number_formats and number_formats.get(i):
            c.number_format = number_formats[i]


def _title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(color=MUTED, size=9, italic=True)


def _sheet_oversigt(wb, plan, dash):
    ws = wb.create_sheet("Oversigt")
    s27 = plan.get("season2027", {})
    wp = s27.get("weightPlan", {})
    goals = plan.get("goals", {})
    zones = plan.get("athletes", {}).get("kennet", {}).get("zones", {})

    _title(ws, "FAST AS FIFTY — status og mål",
           "Genereret fra plan.json og data.json. Redigér aldrig denne fil i hånden.")

    ftp_now = zones.get("ftpW") or zones.get("bikeFtpW")
    w_now = dash.get("coachAssessmentWeightAtGen") or wp.get("startKg")
    f_now = dash.get("coachAssessmentFatAtGen") or wp.get("bodyFatPctStart")
    ctl_now = (dash.get("kpis", {}).get("ctl", {}) or {}).get("value")
    if isinstance(ctl_now, str):
        try:
            ctl_now = float(ctl_now.replace(",", "."))
        except ValueError:
            pass

    def wkg(w, kg):
        try:
            return round(float(w) / float(kg), 2)
        except (TypeError, ValueError, ZeroDivisionError):
            return None

    _headers(ws, 4, ["Måltal", "Nu", "Mål", "Frist", "Note"],
             [26, 12, 12, 14, 62])
    rows = [
        ("Vægt (kg)", w_now, wp.get("targetKg"), wp.get("targetDate"),
         "Max %s kg/uge. Cut starter %s — intet underskud før Médoc."
         % (wp.get("maxLossPerWeekKg"), wp.get("cutStartsFrom"))),
        ("Fedtprocent (Garmin)", f_now, wp.get("bodyFatPctTarget"), wp.get("targetDate"),
         "Bioimpedans — styr efter 7-dages snittet i Tracking, ikke dagstallet."),
        ("FTP (W)", ftp_now, s27.get("ftpTarget"), "2027-08-28",
         "Testporte: okt-26, jan-27, maj-27, jul-27 (Alpe-lejr)."),
        ("W/kg", wkg(ftp_now, w_now), s27.get("wkgTarget"), "2027-08-28",
         "Ca. halvdelen af gevinsten kommer fra vægten, ikke fra watt."),
        ("CTL", ctl_now, s27.get("ctlPeak"), "2027-07-25",
         "Peak %s, %s på startlinjen." % (s27.get("ctlPeak"), s27.get("ctlAtRace"))),
        ("Alkoholfrie dage/uge", None, goals.get("afDaysPerWeek"), "løbende", ""),
        ("Søvn (timer)", None, goals.get("sleepHours"), "løbende", "Ikke forhandleligt."),
    ]
    for i, r in enumerate(rows, start=5):
        _row(ws, i, list(r))

    r = len(rows) + 7
    ws.cell(row=r, column=1, value="LØB").font = GOLD_FONT
    _headers(ws, r + 1, ["Løb", "Dato", "Prioritet", "Distance", "Note"],
             [26, 12, 12, 14, 62])
    races = list(plan.get("races", []))
    races += [
        {"name": "Stelvio Gran Fondo", "date": "2027-06-06", "prio": "B",
         "dist": "", "note": "Hyggetur med Eva + form-check."},
        {"name": "Tour des Stations Ultrafondo", "date": "2027-08-28", "prio": "A",
         "dist": "242 km / 8.848 hm", "note": "Start 02:30. Sæsonens hovedmål."},
    ]
    for i, race in enumerate(races, start=r + 2):
        _row(ws, i, [race.get("name"), race.get("date"), race.get("prio", ""),
                     race.get("dist", ""), race.get("note", "")])
    return ws


def _sheet_plan_2026(wb, plan):
    ws = wb.create_sheet("Plan 2026")
    _title(ws, "Plan 2026 — Christiansborg + Marathon du Médoc",
           "%s uger fra %s" % (plan["program"]["totalWeeks"], plan["program"]["start"]))
    _headers(ws, 4, ["Uge", "Start", "Blok", "CTL-mål", "TSS-mål", "Lokation", "Note"],
             [6, 12, 12, 10, 10, 30, 70])
    for i, w in enumerate(sorted(plan.get("weeks", []), key=lambda x: x["week"]), start=5):
        _row(ws, i, [w.get("week"), w.get("start"), w.get("blockType"),
                     w.get("ctlTarget"), w.get("tssTarget"),
                     w.get("location", ""), w.get("note", "")])
    return ws


def _sheet_plan_2027(wb, plan):
    ws = wb.create_sheet("Plan 2027")
    s27 = plan.get("season2027", {})
    weeks = s27.get("weeks", [])
    if not weeks:
        return ws
    _title(ws, "Plan 2027 — Tour des Stations Ultrafondo",
           "%s uger fra %s til %s · FTP %s → %s W · W/kg %s → %s"
           % (len(weeks), weeks[0]["start"], weeks[-1]["start"],
              s27.get("ftpStart"), s27.get("ftpTarget"),
              s27.get("wkgStart"), s27.get("wkgTarget")))
    _headers(ws, 4, ["Uge", "Start", "Fase", "Blok", "CTL-mål", "FTP-mål",
                     "W/kg-mål", "TSS-mål", "Lokation", "Note"],
             [6, 12, 12, 11, 9, 9, 9, 9, 24, 70])
    fmt = {7: "0.00"}
    for i, w in enumerate(weeks, start=5):
        _row(ws, i, [w.get("week"), w.get("start"), w.get("phase"), w.get("blockType"),
                     w.get("ctlTarget"), w.get("ftpTarget"), w.get("wkgTarget"),
                     w.get("tssTarget"), w.get("location", ""), w.get("note", "")], fmt)
    return ws


def _sheet_tracking(wb, dash):
    ws = wb.create_sheet("Tracking")
    _title(ws, "Tracking — daglig historik",
           "Fra data.json. Vægt og fedt: brug 7-dages snittet som pejling.")
    _headers(ws, 4, ["Dato", "Vægt", "Vægt 7d", "Fedt %", "Fedt 7d",
                     "TSB", "HRV", "Søvn"],
             [12, 10, 10, 10, 10, 10, 10, 10])

    def as_map(key):
        out = {}
        for p in dash.get(key) or []:
            if isinstance(p, dict) and p.get("date") is not None:
                out[p["date"]] = p.get("v")
        return out

    weight, fat = as_map("weightHistory"), as_map("fatHistory")
    tsb, hrv, sleep = as_map("tsbHistory"), as_map("hrvHistory"), as_map("sleepHistory")
    dates = sorted(set(weight) | set(fat) | set(tsb) | set(hrv) | set(sleep))

    def avg7(series, dates_all, idx):
        window = [series.get(d) for d in dates_all[max(0, idx - 6): idx + 1]]
        vals = [v for v in window if isinstance(v, (int, float))]
        return round(sum(vals) / len(vals), 2) if vals else None

    fmt = {2: "0.0", 3: "0.00", 4: "0.0", 5: "0.00", 6: "0.0", 7: "0", 8: "0.0"}
    for i, d in enumerate(dates):
        _row(ws, i + 5, [d, weight.get(d), avg7(weight, dates, i),
                         fat.get(d), avg7(fat, dates, i),
                         tsb.get(d), hrv.get(d), sleep.get(d)], fmt)
    return ws


def _sheet_ugesummer(wb, plan, dash):
    ws = wb.create_sheet("Ugesummer")
    _title(ws, "Ugesummer — planlagt vs. faktisk",
           "Faktisk TSS summeres fra gennemførte sessioner i data.json.")
    _headers(ws, 4, ["Uge", "Start", "Blok", "CTL-mål", "CTL faktisk",
                     "TSS-mål", "TSS faktisk", "Afvigelse"],
             [6, 12, 12, 10, 12, 10, 12, 12])
    ctl_actual = dash.get("ctlCurve") or []
    all_weeks = dash.get("all_weeks") or {}
    for i, w in enumerate(sorted(plan.get("weeks", []), key=lambda x: x["week"]), start=5):
        wk = w["week"]
        actual_ctl = ctl_actual[wk - 1] if wk - 1 < len(ctl_actual) else None
        sess = (all_weeks.get(str(wk)) or {}).get("sessions") or []
        actual_tss = sum(s.get("actual_tss") or 0 for s in sess) or None
        target_tss = w.get("tssTarget")
        dev = (actual_tss - target_tss) if (actual_tss and target_tss) else None
        _row(ws, i, [wk, w.get("start"), w.get("blockType"), w.get("ctlTarget"),
                     actual_ctl, target_tss, actual_tss, dev],
             {5: "0.0", 8: "+0;-0"})
    return ws


def _sheet_zoner(wb, plan):
    ws = wb.create_sheet("Zoner")
    z = plan.get("athletes", {}).get("kennet", {}).get("zones", {})
    _title(ws, "Zoner — aktuelle",
           "Sættes af FTP- og tærskeltest. Opdateres i plan.json, ikke her.")
    _headers(ws, 4, ["Zone", "Løb (pace)", "Cykel (% FTP)", "Cykel (watt)"],
             [10, 20, 20, 20])
    run, bike, watt = z.get("run", {}), z.get("bike", {}), z.get("bikeWatt", {})
    for i, key in enumerate(["Z1", "Z2", "Z3", "Z4", "Z5", "Z6"], start=5):
        _row(ws, i, [key, run.get(key, ""), bike.get(key, ""), watt.get(key, "")])
    r = 12
    _headers(ws, r, ["Nøgletal", "Værdi"], [24, 24])
    for i, (k, v) in enumerate([
        ("Løbetærskel", z.get("runThreshold")),
        ("FTP (W)", z.get("ftpW") or z.get("bikeFtpW")),
    ], start=r + 1):
        _row(ws, i, [k, v])
    return ws


def generate(plan, dash=None):
    """Returnér Kennets Excel-master som bytes."""
    dash = dash or {}
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_oversigt(wb, plan, dash)
    _sheet_plan_2026(wb, plan)
    _sheet_plan_2027(wb, plan)
    _sheet_tracking(wb, dash)
    _sheet_ugesummer(wb, plan, dash)
    _sheet_zoner(wb, plan)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
